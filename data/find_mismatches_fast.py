import sys
import os
import openpyxl

# Add parent to path so we can import sementic
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import sementic

# Fix dataset mappings locally inside sementic
# We removed generic words so now let's see what purely fails
ALL_DATA = sementic.raw_products.get("datasets", {})

test_cases = []
excel_path = os.path.join(os.path.dirname(__file__), "Mospi_Evaluation_Dataset.xlsx")
wb = openpyxl.load_workbook(excel_path)

for prod in ALL_DATA:
    if prod in wb.sheetnames:
        sheet = wb[prod]
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            if not row[0]: continue
            # Limit to 10 queries per product just to identify the theme of failure fast
            if idx > 15: break
            
            prompt = str(row[0]).replace("\n", " ").strip()
            exp_ds = str(row[2] or prod).lower()
            exp_ind = str(row[5] or "").lower()
            test_cases.append((prompt, exp_ds, exp_ind, prod))

print(f"Total test cases selected: {len(test_cases)}")

from sementic import bi_encoder, cross_encoder, faiss_index, INDICATORS, clean_text
import numpy as np

for prompt, expected_ds, expected_ind, original_prod in test_cases:
    # 1. basic fallback expansion locally (replica of route)
    dataset_expansions = {
        r'\bplfs\b': "Periodic Labour Force Survey",
        r'\basuse\b': "Annual Survey of Unincorporated Sector Enterprises",
        r'\basi\b': "Annual Survey of Industries",
        r'\btus\b': "Time Use Survey",
        r'\bgender\b': "Gender Statistics",
        r'\baishe\b': "All India Survey on Higher Education",
        r'\bnfhs\b': "National Family Health Survey",
        r'\bec4\b': "4th Economic Census",
        r'\biip\b': "Index of Industrial Production",
        r'\bwpi\b': "Wholesale Price Index",
        r'\bcpi\b': "Consumer Price Index",
        r'\bnas\b': "National Accounts Statistics",
        r'\brbi\b': "Reserve Bank of India Banking Statistics",
        r'\budise\b': "Unified District Information System for Education Plus"
    }
    q = prompt
    for pat, exp in dataset_expansions.items():
        import re
        if re.search(pat, q.lower()) and exp.lower() not in q.lower():
            q = f"{q} {exp}"
            
    # FORCE RULES
    _force_ds_map = {
        r'\bplfs\b|unemployment rate|labour force|lfpr|wpr|\bworker population ratio\b': ["PLFS"],
        r'\basuse\b|unincorporated|unorganized': ["ASUSE"],
        r'\basi\b|annual survey of industries|factory output|fixed capital|gross output|workers in factory': ["ASI"],
        r'\btus\b|time use survey|unpaid caregiving|domestic services': ["TUS"],
        r'\bgender\b|sex ratio': ["Gender"],
        r'\baishe\b|higher education|college|university': ["AISHE"],
        r'\bcpialrl\b|agricultural labo|rural labo': ["CPIALRL"],
        r'\bhces\b|consumption expenditure|mpce': ["HCES"],
        r'\benvstat\b|environment statistics|forest cover|hazardous waste': ["ENVSTAT"],
        r'\bnfhs\b|family health|immunization|fertility|antenatal care|stunted|wasted|anemia': ["NFHS"],
        r'\bec4\b|4th economic census': ["EC4"],
        r'\bec5\b|5th economic census': ["EC5"],
        r'\bec6\b|6th economic census': ["EC6"],
        r'\biip\b|industrial production|mining index|manufacturing index|electricity index': ["IIP"],
        r'\bwpi\b|wholesale price': ["WPI"],
        r'\bcpi\b|consumer price|retail price|retail inflation': ["CPI", "CPI2"],
        r'\bnas\b|national accounts|gdp|gva': ["NAS"],
        r'\brbi\b|reserve bank|lending rate|exchange rate|forex|external debt|rupee vis-a-vis': ["RBI"],
        r'\bnss79c?\b|cams|modular survey': ["NSS79C", "NSS79"],
        r'\budise\b|school education|unified district': ["UDISE"]
    }
    forced_ds = None
    for pat, ds_list in _force_ds_map.items():
        if re.search(pat, q.lower()):
            forced_ds = ds_list
            break
            
    if forced_ds:
        cands = [i for i in INDICATORS if i["parent"] in forced_ds]
    else:
        q_vec = bi_encoder.encode([clean_text(q)], convert_to_numpy=True)
        q_vec /= np.linalg.norm(q_vec, axis=1, keepdims=True)
        _, I = faiss_index.search(q_vec.astype("float32"), 50)
        cands = [INDICATORS[i] for i in I[0]]
        
    pairs = [(q, c["name"] + " " + c.get("desc", "")) for c in cands]
    scores = cross_encoder.predict(pairs)
    best_idx = int(np.argmax(scores))
    best_cand = cands[best_idx]
    
    pred_ds = best_cand["parent"].lower()
    
    # If the dataset mismatches expected:
    if expected_ds and (expected_ds not in pred_ds) and expected_ds != "none":
        print(f"FAIL_DS | PROXY={original_prod} | PROMPT: {prompt} | PRED_DS: {pred_ds} | PRED_IND: {best_cand['name']} | TARGET_DS: {expected_ds} | REASON: semantic disconnect")
    elif expected_ind and (expected_ind not in best_cand['name'].lower()):
        print(f"FAIL_IND | PROXY={original_prod} | PROMPT: {prompt} | PRED_IND: {best_cand['name']} | TARGET_IND: {expected_ind} | REASON: ind overlap")
