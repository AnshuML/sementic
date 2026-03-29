import openpyxl
import json
import os
import sys

sys.path.append(os.getcwd())
from sementic import app, ESSENTIAL_FILTERS_BY_DATASET

def analyze_errors():
    client = app.test_client()
    report_path = 'data/Semantic Search Test Report_new.xlsx'
    wb = openpyxl.load_workbook(report_path, data_only=True)
    
    # Analyze a few worst performing products
    target_prods = ['PLFS', 'CPI', 'TUS', 'RBI', 'IIP', 'ASI', 'NFHS']
    
    for prod in target_prods:
        print(f"\n{'='*60}\nANALYZING FAILURES FOR: {prod}\n{'='*60}")
        if prod not in wb.sheetnames:
            print(f"Sheet {prod} not found!")
            continue
            
        sheet = wb[prod]
        rows = list(sheet.iter_rows(values_only=True))
        
        for row in rows[1:21]: # up to 20 rows
            prompt = row[0]
            exp_ds = str(row[2] or prod)
            exp_ind = str(row[5] or "")
            if not prompt: continue
            
            try:
                resp = client.post("/search/predict", json={"query": prompt})
                data = resp.get_json()
                
                if not data or "results" not in data or not data["results"]:
                    print(f"[FAIL] Empty result for: {prompt}")
                    continue
                    
                top = data["results"][0]
                pred_ds = top["product"].upper()
                pred_ind = top.get("indicator", "")
                
                is_ds_match = (pred_ds == exp_ds.upper())
                is_ind_match = (exp_ind.lower() in pred_ind.lower() or pred_ind.lower() in exp_ind.lower())
                
                if not is_ds_match or not is_ind_match:
                    print(f"\nPrompt: {prompt}")
                    if not is_ds_match:
                        print(f"  [X] DATASET: Expected '{exp_ds}', Got '{pred_ds}'")
                    if not is_ind_match:
                        print(f"  [X] INDICATOR: Expected '{exp_ind}', Got '{pred_ind}'")
            except Exception as e:
                print(f"Error on prompt {prompt}: {e}")

if __name__ == "__main__":
    analyze_errors()
